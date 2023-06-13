from tempfile import NamedTemporaryFile
from copy import copy

from django.db import transaction
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side

from .models import *
from apps.specifications.services import SpecificationService
from apps.purchases.services import PurchaseService


class ProjectService:
    def __init__(self) -> None:
        self.specification_service = SpecificationService()
        self.purchase_service = PurchaseService()
        
    def _calculate_total_price(self, project: Project) -> int:
        total = 0
        for request in project.specification.request_set.all():
            for offer in request.offer_set.all():
                total += offer.price * offer.count
        return total

    def _get_specification_items(self, project: Project) -> str:
        items = ''
        for request in project.specification.request_set.all():
            for offer in request.offer_set.all():
                items += offer.name + ' ' + str(offer.count) + '\n'
        return items

    def _fit_columns(self, ws: Worksheet) -> None:
        column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
        for column_letter in column_letters:
            ws.column_dimensions[column_letter].auto_size = True
            
    @transaction.atomic
    def create(self, validated_data: dict) -> Project:
        specifcation = validated_data.pop('specification')
        project_obj = Project.objects.create(**validated_data)
        self.specification_service.create(project_obj.id, specifcation)
        self.purchase_service.create(project_obj.id)
        return project_obj
    
    @transaction.atomic
    def update(self, project_id: int, validated_data: dict) -> Project:
        specifcation = validated_data.pop('specification')
        project_obj = Project.objects.get(id=project_id)
        for attribute, value in validated_data:
            setattr(project_obj, attribute, value)
        self.specification_service.update(project_obj.specification_id, specifcation)
        return project_obj

    def export_registration_form(self, project_id: int) -> bytes:
        project_obj = Project.objects.get(id=project_id)
        wb = openpyxl.Workbook()
        headers = ["#", "Дата", "Ответственный за проект", "Наименование субъекта РФ", "Населенный пункт", "Заказчик", "Дилер", "Субдилер",
                   "Наименование продукта, количество", "Сумма поставки", "Сумма аукциона", "Планируемая дата проведения аукциона",
                   "Ссылка на проект на сайте zakupki.gov.ru", "Комментарий", "Номер в базе аукционов"]
        top = Side(border_style='thin')
        bottom = Side(border_style='thin')
        left = Side(border_style='thin')
        right = Side(border_style='thin')

        with NamedTemporaryFile() as tmp:
            ws = wb.active

            # header
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            self._fit_columns(ws)
            ws.row_dimensions[1].height = 100
            ws.row_dimensions[2].height = 100

            # body
            ws.cell(row=2, column=1, value=project_obj.id)
            ws.cell(row=2, column=2, value=str(project_obj.created_at))
            ws.cell(row=2, column=3, value="Чернышев Д.")
            ws.cell(row=2, column=4, value=project_obj.company_region)
            ws.cell(row=2, column=5, value=project_obj.company_city)
            ws.cell(row=2, column=6, value=project_obj.company_name)
            ws.cell(row=2, column=7, value="Кубис")
            ws.cell(row=2, column=8, value=project_obj.partner.name)
            ws.cell(row=2, column=9, value=self._get_specification_items(project_obj))
            ws.cell(row=2, column=10, value=self._calculate_total_price(project_obj))
            ws.cell(row=2, column=11, value="")
            ws.cell(row=2, column=12, value="")
            ws.cell(row=2, column=13, value="")
            ws.cell(row=2, column=14, value="")
            ws.cell(row=2, column=15, value="")
            ws.cell(row=2, column=16, value="")

            for rows in ws.iter_rows(min_row=1, min_col=1):
                for cell in rows:
                    alignment = copy(cell.alignment)
                    alignment.wrapText = True
                    cell.alignment = alignment
                    border = Border(top=top, bottom=bottom, left=left, right=right)
                    cell.border = border

            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream


class ProjectFileService:
    @transaction.atomic
    def create(self, project_id: int, files: list) -> list[File]:
        files_list_response = list()
        for file in files:
                files_list_response.append(File.objects.create(file=file, name=file.name, project_id=project_id))
        return files_list_response