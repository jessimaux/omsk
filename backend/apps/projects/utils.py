from tempfile import NamedTemporaryFile
from copy import copy

import openpyxl
from openpyxl.styles.borders import Border, Side

def get_specification_items(specification):
  total = 0
  for request in specification.request_set.all():
    for offer in request.offer_set.all():
      total += offer.price * offer.count
  return total

def get_total_price(specification):
  items = ''
  for request in specification.request_set.all():
    for offer in request.offer_set.all():
      items += offer.name + ' ' + str(offer.count) + '\n'
  return items

def columns_best_fit(ws):
        """
        Make all columns best fit
        """
        column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
        for column_letter in column_letters:
            ws.column_dimensions[column_letter].auto_size = True

def excel_report(instance):
    wb = openpyxl.Workbook()
    headers = ["#", "Дата", "Ответственный за проект", "Наименование субъекта РФ", "Населенный пункт", "Заказчик", "Дилер", "Субдилер",
               "Наименование продукта, количество", "Сумма поставки", "Сумма аукциона", "Планируемая дата проведения аукциона",
               "Ссылка на проект на сайте zakupki.gov.ru", "Комментарий", "Номер в базе аукционов"]
    top=Side(border_style='thin')
    bottom=Side(border_style='thin')
    left=Side(border_style='thin')
    right=Side(border_style='thin')
    
    with NamedTemporaryFile() as tmp:
        ws = wb.active
        
        # header
        for col, header in enumerate(headers):
            ws.cell(row=1, column=col+1, value=header)
        
        columns_best_fit(ws)
        ws.row_dimensions[1].height = 100
        ws.row_dimensions[2].height = 100
        
        # body
        ws.cell(row=2, column=1, value=instance.id)
        ws.cell(row=2, column=2, value=str(instance.created_at))
        ws.cell(row=2, column=3, value="Чернышев Д.")
        ws.cell(row=2, column=4, value=instance.company_region)
        ws.cell(row=2, column=5, value=instance.company_city)
        ws.cell(row=2, column=6, value=instance.company_name)
        ws.cell(row=2, column=7, value="Кубис")
        ws.cell(row=2, column=8, value=instance.partner.name)
        ws.cell(row=2, column=9, value=get_total_price(instance.specification))
        ws.cell(row=2, column=10, value=get_specification_items(instance.specification))
        ws.cell(row=2, column=11, value="")
        ws.cell(row=2, column=12, value="")
        ws.cell(row=2, column=13, value="")
        ws.cell(row=2, column=14, value="")
        ws.cell(row=2, column=15, value="")
        ws.cell(row=2, column=16, value="")
        
        for rows in ws.iter_rows(min_row=1, min_col=1):
          for cell in rows:
            alignment = copy(cell.alignment)
            alignment.wrapText=True
            cell.alignment = alignment
            border = Border(top=top,bottom=bottom, left=left, right=right)
            cell.border = border
           
        # save as stream
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    return stream