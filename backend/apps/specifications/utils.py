from io import BytesIO
from tempfile import NamedTemporaryFile

from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html  = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result, encoding='utf-8')
    return result.getvalue()


def excel_report(instance, params):
    wb = Workbook()
    headers = ["#", "Наименование по запросу", "Количество", "Цена", "Сумма", "Артикул", "Наименование", 
                 "Количество", "Цена", "Сумма"]
    if(params['str_by_order']=='true'): headers.append('Номер по приказу')
    if(params['link']=='true'): headers.append('Ссылка')
    if(params['country']=='true'): headers.append('Страна происхождения')
    if(params['description']=='true'): headers.append('Описание')
    if(params['description_tech']=='true'): headers.append('ТЗ')
    if(params['description_add']=='true'): headers.append('Заявка')
    
    with NamedTemporaryFile() as tmp:
        ws = wb.active
        
        # header
        ws.merge_cells('A1:E1')
        ws.merge_cells('F1:J1')
        ws['A1'] = 'Запрос'
        ws['F1'] = 'Предложение'
        for col, header in enumerate(headers):
            ws.cell(row=2, column=col+1, value=header)
        
        # body
        row = 3
        row_request = 3
        for no, request in enumerate(instance.request_set.all()):
            row_request = row
            ws.cell(row=row, column=1, value=no+1)
            ws.cell(row=row, column=2, value=request.name)
            ws.cell(row=row, column=3, value=request.amount)
            ws.cell(row=row, column=4, value=request.price)
            ws.cell(row=row, column=5, value=f'=C{row}*D{row}')
            
            for offer in request.offer_set.all():
                ws.cell(row=row, column=6, value=offer.article)
                ws.cell(row=row, column=7, value=offer.name)
                ws.cell(row=row, column=8, value=offer.count)
                ws.cell(row=row, column=9, value=offer.price)
                ws.cell(row=row, column=10, value=f'=C{row_request}*H{row}*I{row}')
                
                col = 11
                if offer.product:
                    if(params['str_by_order']=='true'): 
                        ws.cell(row=row, column=col, value=offer.product.str_by_order)
                        col+=1
                    if(params['link']=='true'): 
                        ws.cell(row=row, column=col, value=offer.product.link)
                        col+=1
                    if(params['country']=='true'): 
                        ws.cell(row=row, column=col, value=offer.product.country)
                        col+=1
                    if(params['description']=='true'): 
                        ws.cell(row=row, column=col, value=offer.product.description)
                        col+=1
                    if(params['description_tech']=='true'): 
                        ws.cell(row=row, column=col, value=offer.product.description_tech)
                        col+=1
                    if(params['description_add']=='true'): 
                        ws.cell(row=row, column=col, value=offer.product.description_add)
                row += 1
         
        ws[f'D{row}'] = 'Итого:'
        ws[f'E{row}'] = f'=SUM(E3:E{row-1})'
        ws[f'I{row}'] = 'Итого:'
        ws[f'J{row}'] = f'=SUM(J3:J{row-1})'
                    
        # save as stream
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    return stream