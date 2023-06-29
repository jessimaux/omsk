from tempfile import NamedTemporaryFile

from django.db import transaction
from openpyxl import Workbook

from .models import *


# TODO: split this with creation service for Request, Offer
class SpecificationService:
    @transaction.atomic
    def create(self, validated_data: dict, project_id: int | None = None) -> Specification | None:
        requests = validated_data.pop('requests')
        specification_obj = Specification.objects.create(project_id=project_id, **validated_data)

        for request in requests:
            offers = request.pop('offers')
            request_obj = Request.objects.create(specification=specification_obj, **request)
            for offer in offers:
                Offer.objects.create(request=request_obj, **offer)

        return specification_obj

    @transaction.atomic
    def update(self, specification_id: int, validated_data: dict) -> Specification | None:
        requests = validated_data.pop('requests')
        requests_from_db = Request.objects.filter(specification_id=specification_id).values_list('id', flat=True)
        requests_id_pool = []

        for request in requests:
            offers = request.pop('offers')
            offers_id_pool = []

            if "id" in request.keys() and Request.objects.filter(id=request['id']).exists():
                request_obj = Request.objects.get(id=request['id'])
                for attribute, value in request.items():
                    setattr(request_obj, attribute, value)
                request_obj.save()
            else:
                request_obj = Request.objects.create(specification_id=specification_id, **request)

            offers_from_db = Offer.objects.filter(request_id=request_obj.id).values_list('id', flat=True)
            requests_id_pool.append(request_obj.id)

            for offer in offers:
                if "id" in offer.keys() and Offer.objects.filter(id=offer['id']).exists():
                    offer_obj = Offer.objects.get(id=offer['id'])
                    for attribute, value in offer.items():
                        setattr(offer_obj, attribute, value)
                    offer_obj.save()
                else:
                    offer_obj = Offer.objects.create(request_id=request_obj.id, **offer)

                offers_id_pool.append(offer_obj.id)

            for offer_id in offers_from_db:
                if offer_id not in offers_id_pool:
                    Offer.objects.filter(id=offer_id).delete()

        for request_id in requests_from_db:
            if request_id not in requests_id_pool:
                Request.objects.filter(id=request_id).delete()

        specification_obj = Specification.objects.get(id=specification_id)
        for attribute, value in validated_data.items():
            setattr(specification_obj, attribute, value)
        specification_obj.save()

        return specification_obj

    def report_xlsx(self, specification_id: int, params: dict) -> bytes:
        specification_obj = Specification.objects.get(id=specification_id)
        wb = Workbook()
        headers = ["#", "Наименование по запросу", "Количество", "Цена", "Сумма", "Артикул", "Наименование",
                   "Количество", "Цена", "Сумма"]
        if (params.get('str_by_order') == 'true'):
            headers.append('Номер по приказу')
        if (params.get('link') == 'true'):
            headers.append('Ссылка')
        if (params.get('country') == 'true'):
            headers.append('Страна происхождения')
        if (params.get('description') == 'true'):
            headers.append('Описание')
        if (params.get('description_tech') == 'true'):
            headers.append('ТЗ')
        if (params.get('description_add') == 'true'):
            headers.append('Заявка')

        with NamedTemporaryFile() as tmp:
            ws = wb.active

            # header
            ws.merge_cells('A1:E1')
            ws.merge_cells('F1:J1')
            ws['A1'] = 'Запрос'
            ws['F1'] = 'Предложение'
            for col, header in enumerate(headers, 1):
                ws.cell(row=2, column=col, value=header)

            # body
            row = 3
            row_request = 3
            for no, request in enumerate(specification_obj.requests.all()):
                row_request = row
                ws.cell(row=row, column=1, value=no+1)
                ws.cell(row=row, column=2, value=request.name)
                ws.cell(row=row, column=3, value=request.amount)
                ws.cell(row=row, column=4, value=request.price)
                ws.cell(row=row, column=5, value=f'=C{row}*D{row}')

                for offer in request.offers.all():
                    ws.cell(row=row, column=6, value=offer.article)
                    ws.cell(row=row, column=7, value=offer.name)
                    ws.cell(row=row, column=8, value=offer.count)
                    ws.cell(row=row, column=9, value=offer.price)
                    ws.cell(row=row, column=10, value=f'=C{row_request}*H{row}*I{row}')

                    col = 11
                    if offer.product:
                        if (params.get('str_by_order') == 'true'):
                            ws.cell(row=row, column=col, value=offer.product.str_by_order)
                            col += 1
                        if (params.get('link') == 'true'):
                            ws.cell(row=row, column=col, value=offer.product.link)
                            col += 1
                        if (params.get('country') == 'true'):
                            ws.cell(row=row, column=col, value=offer.product.country)
                            col += 1
                        if (params.get('description') == 'true'):
                            ws.cell(row=row, column=col, value=offer.product.description)
                            col += 1
                        if (params.get('description_tech') == 'true'):
                            ws.cell(row=row, column=col, value=offer.product.description_tech)
                            col += 1
                        if (params.get('description_add') == 'true'):
                            ws.cell(row=row, column=col, value=offer.product.description_add)
                    row += 1

            ws[f'D{row}'] = 'Итого:'
            ws[f'E{row}'] = f'=SUM(E3:E{row-1})'
            ws[f'I{row}'] = 'Итого:'
            ws[f'J{row}'] = f'=SUM(J3:J{row-1})'

            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream
