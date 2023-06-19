from collections import OrderedDict
from tempfile import NamedTemporaryFile

from django.db import transaction
from django.core.files.uploadedfile import InMemoryUploadedFile
from rest_framework import status
from rest_framework.exceptions import ValidationError
from tablib import Dataset
from openpyxl import Workbook, load_workbook

from .resources import ProductGuideResource
from .models import *

# TODO: split services with contact by create contact service


class ProductService:
    @transaction.atomic
    def import_xlsx(self, uploaded_file: InMemoryUploadedFile) -> OrderedDict:
        product_resource = ProductGuideResource()
        dataset = Dataset()
        dataset.load(uploaded_file)
        result = product_resource.import_data(dataset)
        if result.has_errors():
            raise ValidationError({'non_field_errors': 'Ошибка импорта.'}, code=status.HTTP_400_BAD_REQUEST)
        return result.totals


class PartnerService:
    ID = 0
    NAME = 1
    INN = 2
    REGION = 3
    DISCOUNT = 4
    CONTACT_ID = 5
    CONTACT_PARTNER_ID = 6
    CONTACT_FIO = 7
    CONTACT_ROLE = 8
    CONTACT_PHONE = 9
    CONTACT_EMAIL = 10

    def _contact_partner_handle(self, row: list, partner_id: int) -> None:
        contact_fields = ['partner_id', 'fio', 'role', 'phone', 'email']

        if row[self.CONTACT_ID]:
            contact_obj = ContactPartner.objects.get(id=row[self.CONTACT_ID])
            for index, attribute in enumerate(contact_fields, self.CONTACT_PARTNER_ID):
                setattr(contact_obj, attribute, row[index])
            contact_obj.save()
        else:
            ContactPartner.objects.create(partner_id=partner_id,
                                          fio=row[self.CONTACT_FIO],
                                          role=row[self.CONTACT_ROLE],
                                          phone=row[self.CONTACT_PHONE],
                                          email=row[self.CONTACT_EMAIL])

    def _is_row_partner_handler(self, row: list) -> bool:
        for attr in row[self.ID:self.DISCOUNT]:
            if attr:
                return True

    @transaction.atomic
    def create(self, validated_data: dict) -> PartnerGuide | None:
        contacts = validated_data.pop('contact_partner')
        partner_obj = PartnerGuide.objects.create(**validated_data)
        for contact in contacts:
            ContactPartner.objects.create(partner_id=partner_obj.id, **contact)
        return partner_obj

    @transaction.atomic
    def update(self, partner_id: int, validated_data: dict) -> PartnerGuide | None:
        contacts = validated_data.pop('contact_partner')
        contacts_from_db = ContactPartner.objects.filter(partner_id=partner_id).values_list('id', flat=True)
        contacts_id_pool = []

        for contact in contacts:
            if "id" in contact.keys() and ContactPartner.objects.filter(id=contact['id']).exists():
                contact_obj = ContactPartner.objects.get(id=contact['id'])
                for attribute, value in contact.items():
                    setattr(contact_obj, attribute, value)
                contact_obj.save()
            else:
                contact_obj = ContactPartner.objects.create(partner_id=partner_id, **contact)

            contacts_id_pool.append(contact_obj.id)

        for contact_id in contacts_from_db:
            if contact_id not in contacts_id_pool:
                ContactPartner.objects.filter(id=contact_id).delete()

        partner_obj = PartnerGuide.objects.get(id=partner_id)
        for attribute, value in validated_data.items():
            setattr(partner_obj, attribute, value)
        partner_obj.save()

        return partner_obj

    def export_xlsx(self) -> bytes:
        wb = Workbook()
        PARTNERS_FIELDS = ['id', 'name', 'inn', 'region', 'discount']
        LEN_PARTNERS_FIELDS = len(PARTNERS_FIELDS)
        CONTACT_FIELDS = ['id', 'partner_id', 'fio', 'role', 'phone', 'email']
        partners_query = PartnerGuide.objects.all()

        with NamedTemporaryFile() as tmp:
            ws = wb.active
            # fill header
            for col, title in enumerate(PARTNERS_FIELDS+CONTACT_FIELDS, 1):
                ws.cell(row=1, column=col, value=title)

            # fill body
            row = 2
            for partner in partners_query:
                # fill information about partner
                for col, attr in enumerate(PARTNERS_FIELDS, 1):
                    ws.cell(row=row, column=col, value=getattr(partner, attr))

                # fill information about partner's contacts
                for contact in partner.contact_partner.all():
                    for col, attr in enumerate(CONTACT_FIELDS, 1):
                        ws.cell(row=row, column=LEN_PARTNERS_FIELDS+col, value=getattr(contact, attr))
                    row += 1

            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream

    @transaction.atomic
    def import_xlsx(self, upload_file: InMemoryUploadedFile) -> None:
        wb = load_workbook(upload_file)
        ws = wb.active
        partners_fields = ['name', 'inn', 'region', 'discount']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if self._is_row_partner_handler(row):
                if row[self.ID]:
                    partner_obj = PartnerGuide.objects.get(id=row[self.ID])
                    for index, attr in enumerate(partners_fields, self.NAME):
                        setattr(partner_obj, attr, row[index])
                    partner_obj.save()
                    self._contact_partner_handle(row, partner_obj.id)
                else:
                    partner_obj = PartnerGuide.objects.create(name=row[self.NAME],
                                                              inn=row[self.INN],
                                                              region=row[self.REGION],
                                                              discount=row[self.DISCOUNT])
                    self._contact_partner_handle(row, partner_obj.id)
            else:
                self._contact_partner_handle(row, partner_obj.id)


class ProviderService:
    ID = 0
    NAME = 1
    SPHERE = 2
    INN = 3
    REGION = 4
    DISCOUNT = 5
    CONTACT_ID = 6
    CONTACT_PROVIDER_ID = 7
    CONTACT_FIO = 8
    CONTACT_ROLE = 9
    CONTACT_PHONE = 10
    CONTACT_EMAIL = 11

    def _is_row_provider_handler(self, row: list) -> bool:
        for attr in row[self.ID:self.DISCOUNT]:
            if attr:
                return True

    def _contact_provider_handle(self, row: list, provider_id: int) -> None:
        contact_fields = ['provider_id', 'fio', 'role', 'phone', 'email']

        if row[self.CONTACT_PROVIDER_ID]:
            contact_obj = ContactProvider.objects.get(id=row[self.CONTACT_PROVIDER_ID])
            for index, attribute in enumerate(contact_fields, self.CONTACT_PROVIDER_ID):
                setattr(contact_obj, attribute, row[index])
            contact_obj.save()
        else:
            ContactProvider.objects.create(provider_id=provider_id,
                                           fio=row[self.CONTACT_FIO],
                                           role=row[self.CONTACT_ROLE],
                                           phone=row[self.CONTACT_PHONE],
                                           email=row[self.CONTACT_EMAIL])

    @transaction.atomic
    def create(self, validated_data: dict) -> ProviderGuide | None:
        contacts = validated_data.pop('contact_provider')
        provider_obj = ProviderGuide.objects.create(**validated_data)
        for contact in contacts:
            ContactProvider.objects.create(provider_id=provider_obj.id, **contact)
        return provider_obj

    @transaction.atomic
    def update(self, provider_id: int, validated_data: dict) -> ProviderGuide | None:
        contacts = validated_data.pop('contact_provider')
        contacts_from_db = ContactProvider.objects.filter(provider_id=provider_id).values_list('id', flat=True)
        contacts_id_pool = []

        for contact in contacts:
            if "id" in contact.keys() and ContactProvider.objects.filter(id=contact['id']).exists():
                contact_obj = ContactProvider.objects.get(id=contact['id'])
                for attribute, value in contact.items():
                    setattr(contact_obj, attribute, value)
                contact_obj.save()
            else:
                contact_obj = ContactProvider.objects.create(provider_id=provider_id, **contact)

            contacts_id_pool.append(contact_obj.id)

        for contact_id in contacts_from_db:
            if contact_id not in contacts_id_pool:
                ContactProvider.objects.filter(id=contact_id).delete()

        provider_obj = ProviderGuide.objects.get(id=provider_id)
        for attribute, value in validated_data.items():
            setattr(provider_obj, attribute, value)
        provider_obj.save()

        return provider_obj

    def export_xlsx(self) -> bytes:
        wb = Workbook()
        PROVIDERS_FIELDS = ['id', 'name', 'sphere', 'inn', 'region', 'discount']
        LEN_PROVIDERS_FIELDS = len(PROVIDERS_FIELDS)
        CONTACT_FIELDS = ['id', 'provider_id', 'fio', 'role', 'phone', 'email']
        providers_query = ProviderGuide.objects.all()

        with NamedTemporaryFile() as tmp:
            ws = wb.active
            # fill header
            for col, title in enumerate(PROVIDERS_FIELDS+CONTACT_FIELDS, 1):
                ws.cell(row=1, column=col, value=title)

            # fill body
            row = 2
            for provider in providers_query:
                # fill information about provider
                for col, attr in enumerate(PROVIDERS_FIELDS, 1):
                    ws.cell(row=row, column=col, value=getattr(provider, attr))

                # fill information about provider's contacts
                for contact in provider.contact_provider.all():
                    for col, attr in enumerate(CONTACT_FIELDS, 1):
                        ws.cell(row=row, column=LEN_PROVIDERS_FIELDS+col, value=getattr(contact, attr))
                    row += 1

            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream

    @transaction.atomic
    def import_xlsx(self, upload_file: InMemoryUploadedFile) -> None:
        wb = load_workbook(upload_file)
        ws = wb.active
        providers_fields = ['name', 'sphere', 'inn', 'region', 'discount']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if self._is_row_provider_handler(row):
                if row[self.ID]:
                    provider_obj = ProviderGuide.objects.get(id=row[self.ID])
                    for index, attr in enumerate(providers_fields, self.NAME):
                        setattr(provider_obj, attr, row[index])
                    provider_obj.save()
                    self._contact_provider_handle(row, provider_obj.id)
                else:
                    provider_obj = ProviderGuide.objects.create(name=row[self.NAME],
                                                                sphere=row[self.SPHERE],
                                                                inn=row[self.INN],
                                                                region=row[self.REGION],
                                                                discount=row[self.DISCOUNT])
                    self._contact_provider_handle(row, provider_obj.id)
            else:
                self._contact_provider_handle(row, provider_obj.id)
