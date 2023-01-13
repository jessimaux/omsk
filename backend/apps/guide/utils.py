from tempfile import NamedTemporaryFile

from django.db import transaction
from rest_framework import status
from rest_framework.response import Response
from openpyxl import Workbook, load_workbook

from .models import PartnerGuide, ProviderGuide, ContactPartner, ContactProvider


def check_xlsx_file_import(request, serializer_class):
    """
    Check file on exists and correct format
    """
    if 'file' not in request.FILES or not serializer_class.is_valid():
        return Response({'non_field_errors': 'Файл не выбран.'}, status=status.HTTP_400_BAD_REQUEST)
    elif request.FILES['file'].content_type not in ('application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
        return Response({'non_field_errors': 'Выбран некорректный формат файла.'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        return True

def contact_partner_handle(row, partner_id):
    """
    Method to handle contact row from xlsx import
    """
    contact_fields = ['partner_id', 'fio', 'role', 'phone', 'email']
    
    if row[0]:
        contact_obj = ContactPartner.objects.get(id=row[0])
        for index, attribute in enumerate(contact_fields):
            setattr(contact_obj, attribute, row[index+1])
        contact_obj.save()
    else:
        ContactPartner.objects.create(partner_id=partner_id, fio=row[2], role=row[3], phone=row[4], email=row[5])

def import_partners(upload_file):
    wb = load_workbook(upload_file)
    ws = wb.active
    
    partners_fields = ['name', 'inn', 'region', 'discount']
    
    # if something went wrong transacation will rollback
    with transaction.atomic():
        partner_id = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            # check row for empty partner fields: if flag - handle partner with contact, else handle only contact
            flag_empty = False
            for attr in row[0:4]:
                if attr:
                    flag_empty = True
                    break
           
            if flag_empty:
                if row[0]:
                    partner_obj = PartnerGuide.objects.get(id=row[0])
                    partner_id = row[0]
                    for index, attr in enumerate(partners_fields):
                        setattr(partner_obj, attr, row[index+1])
                    partner_obj.save()
                    contact_partner_handle(row[5:11], partner_id)
                else:
                    partner_obj = PartnerGuide.objects.create(name=row[1], inn=row[2], region=row[3], discount=row[4])
                    partner_id = partner_obj.id
                    contact_partner_handle(row[5:11], partner_id)
            else:
                contact_partner_handle(row[5:11], partner_id)
            
def export_partners():
    wb = Workbook()
    
    partners_query = PartnerGuide.objects.all()
    
    partners_fields = ['id', 'name', 'inn', 'region', 'discount']
    contact_fields = ['id', 'partner_id', 'fio', 'role', 'phone', 'email']
        
    with NamedTemporaryFile() as tmp:
        ws = wb.active
        
        # header
        for col, title in enumerate(partners_fields+contact_fields):
            ws.cell(row=1, column=col+1, value=title)
        
        # body
        row = 2
        for partner in partners_query:
            for col, attr in enumerate(partners_fields):
                ws.cell(row=row, column=col+1, value=getattr(partner, attr))
            for contact in partner.contact_partner.all():
                for col, attr in enumerate(contact_fields):
                    ws.cell(row=row, column=len(partners_fields)+col+1, value=getattr(contact, attr))
                row += 1
                    
        # save as stream
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    return stream


def contact_provider_handle(row, provider_id):
    """
    Method to handle contact row from xlsx import
    """
    contact_fields = ['provider_id', 'fio', 'role', 'phone', 'email']
    
    if row[0]:
        contact_obj = ContactProvider.objects.get(id=row[0])
        for index, attribute in enumerate(contact_fields):
            setattr(contact_obj, attribute, row[index+1])
        contact_obj.save()
    else:
        ContactProvider.objects.create(provider_id=provider_id, fio=row[2], role=row[3], phone=row[4], email=row[5])

def import_providers(upload_file):
    wb = load_workbook(upload_file)
    ws = wb.active
    
    providers_fields = ['sphere', 'name', 'inn', 'region', 'discount']
    
    # if something went wrong transacation will rollback
    with transaction.atomic():
        provider_id = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            # check for empty provider row
            flag_empty = False
            for attr in row[0:5]:
                if attr:
                    flag_empty = True
                    break
           
            # check row for empty partner fields: if flag - handle partner with contact, else handle only contact
            if flag_empty:
                if row[0]:
                    provider_obj = ProviderGuide.objects.get(id=row[0])
                    provider_id = row[0]
                    for index, attr in enumerate(providers_fields):
                        setattr(provider_obj, attr, row[index+1])
                    provider_obj.save()
                    contact_provider_handle(row[6:12], provider_id)
                else:
                    provider_obj = ProviderGuide.objects.create(sphere=row[1], name=row[2], inn=row[3], region=row[4], discount=row[5])
                    provider_id = provider_obj.id
                    contact_provider_handle(row[6:12], provider_id)
            else:
                contact_provider_handle(row[6:12], provider_id)
            
def export_providers():
    wb = Workbook()
    
    providers_query = ProviderGuide.objects.all()
    
    providers_fields = ['id', 'sphere', 'name', 'inn', 'region', 'discount']
    contact_fields = ['id', 'provider_id', 'fio', 'role', 'phone', 'email']
        
    with NamedTemporaryFile() as tmp:
        ws = wb.active
        
        # header
        for col, title in enumerate(providers_fields+contact_fields):
            ws.cell(row=1, column=col+1, value=title)
        
        # body
        row = 2
        for provider in providers_query:
            for col, attr in enumerate(providers_fields):
                ws.cell(row=row, column=col+1, value=getattr(provider, attr))
            for contact in provider.contact_provider.all():
                for col, attr in enumerate(contact_fields):
                    ws.cell(row=row, column=len(providers_fields)+col+1, value=getattr(contact, attr))
                row += 1
                    
        # save as stream
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    return stream