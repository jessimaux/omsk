from tempfile import NamedTemporaryFile

from django.db import transaction
from openpyxl import Workbook

from .models import *


class PurchaseService:
    @transaction.atomic
    def create(self, project_id: int) -> Purchase:
        purchase_obj = Purchase.objects.create(project_id=project_id)
        for request in purchase_obj.project.specification.requests:
            for offer in request.offers:
                PurchaseOffer.objects.create(purchase_id=purchase_obj.id, 
                                            offer_id=offer.id,
                                            status='Заказан',
                                            price_buy=offer.product.price_buy if offer.product else 0,
                                            nds_base=offer.product.nds if offer.product else 0)
        
    
    @transaction.atomic
    def update(self, purchase_id: int, validated_data: dict) -> Purchase:
        purchases = validated_data.pop('purchases')
        
        purchase_obj = Purchase.objects.get(id=purchase_id)
        for attribute, value in validated_data:
            setattr(purchase_obj, attribute, value)
        
        for purchase_item in purchases:
            if PurchaseOffer.objects.filter(id=purchase_item['id']):
                purchase_item_obj = PurchaseOffer.objects.get(id=purchase_item['id'])
                for attr, value in purchase_item.items():
                    setattr(purchase_item_obj, attr, value)
                purchase_item_obj.save()
        return purchase_obj
    
    
    def export_xlsx(self, purchase_id: int) -> bytes:
        purchase_obj = Purchase.objects.get(id=purchase_id)
        wb = Workbook()
        headers = ["#", "Статус товара", "ISBN поставщика", "Наименование спецификации", "Страна происхождения по спецификации", "Страна происхождения по базе", 
                "Количество", "Артикул", "Наименование в базе", "Количество в 1 ковмплекте", "Поставщик", "Цена закупки", "Сумма закупки", "Цена продажи",
                "Сумма продажи", "НДС(база)", "НДС(продажа)", "Рентабельность %", "Рентабельность РУБ", "Спрок поставки", "Предоплата", "Входящие счета",
                "Оплата входяшего счета", "Ориентировочный срок прихода на склад"]
        with NamedTemporaryFile() as tmp:
            ws = wb.active
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            row = 2
            for no, purchase in enumerate(purchase_obj.purchases.all()):
                ws.cell(row=row, column=1, value=no+1)
                ws.cell(row=row, column=2, value=purchase.status)
                ws.cell(row=row, column=3, value=purchase.isbn)
                ws.cell(row=row, column=4, value=purchase.offer.name)
                ws.cell(row=row, column=5, value=purchase.country)
                if purchase.offer.product:
                    ws.cell(row=row, column=6, value=purchase.offer.product.country)
                    ws.cell(row=row, column=9, value=purchase.offer.product.name)
                ws.cell(row=row, column=7, value=purchase.offer.request.amount)
                ws.cell(row=row, column=8, value=purchase.offer.article)
                ws.cell(row=row, column=10, value=purchase.offer.count)
                ws.cell(row=row, column=11, value='?')
                ws.cell(row=row, column=12, value=purchase.price_buy)
                ws.cell(row=row, column=13, value=purchase.offer.request.amount*purchase.offer.count*purchase.price_buy)
                ws.cell(row=row, column=14, value=purchase.offer.price)
                ws.cell(row=row, column=15, value=purchase.offer.request.amount*purchase.offer.count*purchase.offer.price)
                ws.cell(row=row, column=16, value=purchase.nds_base)
                ws.cell(row=row, column=17, value=purchase.nds_sell)
                try:
                    ws.cell(row=row, column=18, value=1 - purchase.price_buy/purchase.offer.price)
                except ZeroDivisionError:
                    ws.cell(row=row, column=18, value='inf')
                ws.cell(row=row, column=19, value=purchase.offer.request.amount*purchase.offer.count*(purchase.price_buy-purchase.offer.price))
                ws.cell(row=row, column=20, value=purchase.delivery_period)
                ws.cell(row=row, column=21, value=purchase.prepayment)
                ws.cell(row=row, column=22, value=purchase.bill_income)
                ws.cell(row=row, column=23, value=purchase.bill_income_complete)
                ws.cell(row=row, column=24, value=purchase.warehouse_delivery_date)
                
                row += 1

            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return stream